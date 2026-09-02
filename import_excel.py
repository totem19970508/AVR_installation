from __future__ import annotations

import argparse
import math
from collections.abc import Iterator, Sequence
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.firebase_client import get_firestore_client
from src.models import SkylightInstallation

COLLECTION_NAME = "skylight_installation"
EXPECTED_HEADERS = (
    "Plan No",
    "REGION",
    "ROW",
    "POSITION",
    "Measured Dmensions",
    "X-bar coverage size",
    "Pixels",
    "Actual length",
    2000,
    1500,
    1000,
    "Cutting length",
    "Cutted Pixel",
    "Actual Cutted Pixel",
    "Remarks",
)


def _text(value: Any, field: str) -> str:
    if value is None:
        raise ValueError(f"{field} is blank")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _number(value: Any, field: str) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"{field} must be numeric, got {value!r}")
    return float(value)


def _optional_number(value: Any, field: str) -> float | None:
    return None if value is None else _number(value, field)


def _integer(value: Any, field: str, *, optional: bool = False) -> int | None:
    if value is None and optional:
        return None
    number = _number(value, field)
    rounded = round(number)
    if not math.isclose(number, rounded, rel_tol=0.0, abs_tol=1e-9):
        raise ValueError(f"{field} must be a whole number, got {value!r}")
    return rounded


def _document_id(record: SkylightInstallation) -> str:
    return f"installation-{record.plan_no}"


def _parse_row(values: Sequence[Any], excel_row: int) -> SkylightInstallation:
    try:
        return SkylightInstallation(
            plan_no=_text(values[0], "Plan No"),
            region=_text(values[1], "REGION"),
            row=_text(values[2], "ROW"),
            position=_text(values[3], "POSITION"),
            measured_dimensions=_number(values[4], "Measured Dmensions"),
            xbar_coverage_size=_number(values[5], "X-bar coverage size"),
            pixels=_integer(values[6], "Pixels"),
            actual_length=_number(values[7], "Actual length"),
            progress={
                "2000_mm": _integer(values[8], "2000", optional=True),
                "1500_mm": _integer(values[9], "1500", optional=True),
                "1000_mm": _integer(values[10], "1000", optional=True),
            },
            cutting_length=_optional_number(values[11], "Cutting length"),
            cutted_pixel=_integer(values[12], "Cutted Pixel"),
            actual_cutted_pixel=_integer(
                values[13], "Actual Cutted Pixel", optional=True
            ),
            remarks=None if values[14] is None else str(values[14]).strip(),
        )
    except (TypeError, ValueError) as error:
        raise ValueError(f"Excel row {excel_row}: {error}") from error


def read_installations(path: Path) -> Iterator[tuple[str, SkylightInstallation]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = tuple(
        cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
    )
    if headers != EXPECTED_HEADERS:
        raise ValueError(f"Unexpected headers in {worksheet.title}: {headers!r}")

    seen_ids: set[str] = set()
    for excel_row, values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if all(values[index] is None for index in range(4)):
            continue

        record = _parse_row(values, excel_row)
        document_id = _document_id(record)
        if document_id in seen_ids:
            raise ValueError(f"Excel row {excel_row}: duplicate installation key")
        seen_ids.add(document_id)
        yield document_id, record


def upload(records: list[tuple[str, SkylightInstallation]]) -> None:
    database = get_firestore_client()
    collection = database.collection(COLLECTION_NAME)

    for offset in range(0, len(records), 450):
        batch = database.batch()
        for document_id, record in records[offset : offset + 450]:
            document = record.to_firestore()
            if document["actual_cutted_pixel"] is None:
                del document["actual_cutted_pixel"]
            if document["remarks"] is None:
                del document["remarks"]
            batch.set(
                collection.document(document_id), document, merge=True
            )
        batch.commit()


def main() -> None:
    parser = argparse.ArgumentParser(description="Import skylight installations")
    parser.add_argument("workbook", nargs="?", default="data_base.xlsx", type=Path)
    parser.add_argument(
        "--commit",
        action="store_true",
        help="write validated records to Firestore; otherwise only preview",
    )
    arguments = parser.parse_args()

    records = list(read_installations(arguments.workbook))
    action = "Uploading" if arguments.commit else "Previewed"
    print(f"{action} {len(records)} records for {COLLECTION_NAME}")
    print("First document IDs:", ", ".join(item[0] for item in records[:3]))

    if arguments.commit:
        upload(records)
        print(f"Uploaded {len(records)} records successfully")
    else:
        print("No Firebase changes made. Add --commit to upload.")


if __name__ == "__main__":
    main()
